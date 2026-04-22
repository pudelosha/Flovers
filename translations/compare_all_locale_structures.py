import json
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import config


LOCALE_ROOTS = [
    Path(r"C:\Projekty\Python\Flovers\backend\i18n\locales"),
    Path(r"C:\Projekty\Python\Flovers\mobile\src\i18n\locales"),
    Path(r"C:\Projekty\Python\Flovers\web\src\locales"),
]


@dataclass
class CompareResult:
    root_dir: Path
    rel_path: Path
    lang: str
    status: str  # OK / MISSING / INVALID / DIFF / EXTRA / ERROR
    file_path: Path
    message: str = ""
    diffs: List[str] = field(default_factory=list)


def read_json_file(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def format_path(path: str) -> str:
    return path if path else "$"


def compare_structures(source: Any, target: Any, path: str = "") -> List[str]:
    diffs: List[str] = []

    if type(source) is not type(target):
        diffs.append(
            f"{format_path(path)}: type mismatch - source={type(source).__name__}, target={type(target).__name__}"
        )
        return diffs

    if isinstance(source, dict):
        source_keys = list(source.keys())
        target_keys = list(target.keys())

        if source_keys != target_keys:
            source_key_set = set(source_keys)
            target_key_set = set(target_keys)

            missing = [k for k in source_keys if k not in target_key_set]
            extra = [k for k in target_keys if k not in source_key_set]

            if missing:
                diffs.append(f"{format_path(path)}: missing keys in target: {missing}")

            if extra:
                diffs.append(f"{format_path(path)}: extra keys in target: {extra}")

            common_source_order = [k for k in source_keys if k in target_key_set]
            common_target_order = [k for k in target_keys if k in source_key_set]
            if common_source_order != common_target_order:
                diffs.append(f"{format_path(path)}: key order mismatch")

        for key in source_keys:
            if key in target:
                child_path = f"{path}.{key}" if path else key
                diffs.extend(compare_structures(source[key], target[key], child_path))

        return diffs

    if isinstance(source, list):
        if len(source) != len(target):
            diffs.append(
                f"{format_path(path)}: list length mismatch - source={len(source)}, target={len(target)}"
            )
            return diffs

        for i, (s_item, t_item) in enumerate(zip(source, target)):
            child_path = f"{path}[{i}]" if path else f"[{i}]"
            diffs.extend(compare_structures(s_item, t_item, child_path))

        return diffs

    return diffs


def validate_root_dir(root_dir: Path) -> None:
    if not root_dir.exists():
        raise FileNotFoundError(f"Locale root does not exist: {root_dir}")

    if not root_dir.is_dir():
        raise ValueError(f"Locale root is not a directory: {root_dir}")

    source_dir = root_dir / config.SOURCE_LANG
    if not source_dir.exists():
        raise FileNotFoundError(
            f"Source language folder '{config.SOURCE_LANG}' not found in: {root_dir}"
        )

    if not source_dir.is_dir():
        raise ValueError(f"Source language path is not a directory: {source_dir}")


def collect_json_files_by_rel_path(lang_dir: Path) -> Dict[Path, Path]:
    files: Dict[Path, Path] = {}

    for path in lang_dir.rglob("*.json"):
        if path.is_file():
            rel_path = path.relative_to(lang_dir)
            files[rel_path] = path

    return files


def compare_one_file(
    root_dir: Path,
    rel_path: Path,
    source_file: Path,
    target_lang: str,
) -> CompareResult:
    target_file = root_dir / target_lang / rel_path

    if not target_file.exists():
        return CompareResult(
            root_dir=root_dir,
            rel_path=rel_path,
            lang=target_lang,
            status="MISSING",
            file_path=target_file,
            message="Target file does not exist",
        )

    if not target_file.is_file():
        return CompareResult(
            root_dir=root_dir,
            rel_path=rel_path,
            lang=target_lang,
            status="ERROR",
            file_path=target_file,
            message="Target path exists but is not a file",
        )

    try:
        source_data = read_json_file(source_file)
    except Exception as e:
        return CompareResult(
            root_dir=root_dir,
            rel_path=rel_path,
            lang=target_lang,
            status="ERROR",
            file_path=source_file,
            message=f"Source JSON invalid: {str(e)[:200]}",
        )

    try:
        target_data = read_json_file(target_file)
    except Exception as e:
        return CompareResult(
            root_dir=root_dir,
            rel_path=rel_path,
            lang=target_lang,
            status="INVALID",
            file_path=target_file,
            message=f"Invalid JSON: {str(e)[:200]}",
        )

    diffs = compare_structures(source_data, target_data)

    if diffs:
        return CompareResult(
            root_dir=root_dir,
            rel_path=rel_path,
            lang=target_lang,
            status="DIFF",
            file_path=target_file,
            message=f"{len(diffs)} structural difference(s) found",
            diffs=diffs,
        )

    return CompareResult(
        root_dir=root_dir,
        rel_path=rel_path,
        lang=target_lang,
        status="OK",
        file_path=target_file,
        message="Structure matches",
    )


def find_extra_target_files(
    root_dir: Path,
    target_lang: str,
    source_rel_paths: Set[Path],
) -> List[CompareResult]:
    results: List[CompareResult] = []
    target_dir = root_dir / target_lang

    if not target_dir.exists():
        return results

    if not target_dir.is_dir():
        results.append(
            CompareResult(
                root_dir=root_dir,
                rel_path=Path(),
                lang=target_lang,
                status="ERROR",
                file_path=target_dir,
                message="Target language path exists but is not a directory",
            )
        )
        return results

    target_files = collect_json_files_by_rel_path(target_dir)

    for rel_path, file_path in sorted(target_files.items()):
        if rel_path not in source_rel_paths:
            results.append(
                CompareResult(
                    root_dir=root_dir,
                    rel_path=rel_path,
                    lang=target_lang,
                    status="EXTRA",
                    file_path=file_path,
                    message="File exists in target language but not in source EN",
                )
            )

    return results


def process_root(root_dir: Path) -> Tuple[List[CompareResult], Dict[str, int]]:
    validate_root_dir(root_dir)

    source_dir = root_dir / config.SOURCE_LANG
    source_files = collect_json_files_by_rel_path(source_dir)
    source_rel_paths = set(source_files.keys())

    results: List[CompareResult] = []
    counters = {"OK": 0, "MISSING": 0, "INVALID": 0, "DIFF": 0, "EXTRA": 0, "ERROR": 0}

    for rel_path, source_file in sorted(source_files.items()):
        for lang in config.TARGET_LANGS:
            result = compare_one_file(root_dir, rel_path, source_file, lang)
            results.append(result)
            counters[result.status] = counters.get(result.status, 0) + 1

    for lang in config.TARGET_LANGS:
        extra_results = find_extra_target_files(root_dir, lang, source_rel_paths)
        for result in extra_results:
            results.append(result)
            counters[result.status] = counters.get(result.status, 0) + 1

    return results, counters


def auto_fit_worksheet(ws) -> None:
    max_widths: Dict[int, int] = {}

    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            width = min(len(value) + 2, 100)
            col_idx = cell.column
            max_widths[col_idx] = max(max_widths.get(col_idx, 0), width)

    for col_idx, width in max_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def style_status_cell(cell) -> None:
    fills = {
        "OK": "C6EFCE",
        "MISSING": "FFC7CE",
        "INVALID": "FFEB9C",
        "DIFF": "F4CCCC",
        "EXTRA": "D9EAD3",
        "ERROR": "EA9999",
    }
    color = fills.get(str(cell.value), None)
    if color:
        cell.fill = PatternFill(fill_type="solid", start_color=color, end_color=color)
        cell.font = Font(bold=True)


def write_summary_sheet(
    wb: Workbook,
    per_root_summary: List[Tuple[Path, Dict[str, int]]],
    total_summary: Dict[str, int],
) -> None:
    ws = wb.active
    ws.title = "Summary"

    headers = ["Root", "OK", "MISSING", "INVALID", "DIFF", "EXTRA", "ERROR"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")

    for root_dir, counters in per_root_summary:
        ws.append(
            [
                str(root_dir),
                counters["OK"],
                counters["MISSING"],
                counters["INVALID"],
                counters["DIFF"],
                counters["EXTRA"],
                counters["ERROR"],
            ]
        )

    ws.append([])
    ws.append(
        [
            "TOTAL",
            total_summary["OK"],
            total_summary["MISSING"],
            total_summary["INVALID"],
            total_summary["DIFF"],
            total_summary["EXTRA"],
            total_summary["ERROR"],
        ]
    )

    total_row = ws.max_row
    for cell in ws[total_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", start_color="BDD7EE", end_color="BDD7EE")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"
    auto_fit_worksheet(ws)


def write_details_sheet(wb: Workbook, results: List[CompareResult]) -> None:
    ws = wb.create_sheet("Details")

    headers = [
        "Root",
        "Language",
        "Relative Path",
        "Status",
        "Message",
        "File Path",
        "Diff Count",
        "Diffs",
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")

    for result in results:
        ws.append(
            [
                str(result.root_dir),
                result.lang,
                result.rel_path.as_posix(),
                result.status,
                result.message,
                str(result.file_path),
                len(result.diffs),
                "\n".join(result.diffs),
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        style_status_cell(row[3])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[7].alignment = row[7].alignment.copy(wrap_text=True)

    auto_fit_worksheet(ws)
    ws.column_dimensions["H"].width = 100


def build_report(results: List[CompareResult], per_root_summary: List[Tuple[Path, Dict[str, int]]]) -> Path:
    total_summary = {"OK": 0, "MISSING": 0, "INVALID": 0, "DIFF": 0, "EXTRA": 0, "ERROR": 0}

    for _, counters in per_root_summary:
        for key in total_summary:
            total_summary[key] += counters.get(key, 0)

    wb = Workbook()
    write_summary_sheet(wb, per_root_summary, total_summary)
    write_details_sheet(wb, results)

    output_dir = Path(__file__).parent
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"locale_structure_report_{timestamp}.xlsx"
    wb.save(output_path)

    return output_path


def main() -> None:
    all_results: List[CompareResult] = []
    per_root_summary: List[Tuple[Path, Dict[str, int]]] = []

    print(f"Configured source language: {config.SOURCE_LANG}")
    print(f"Configured languages: {', '.join(config.LANGS)}")
    print()

    for root_dir in LOCALE_ROOTS:
        try:
            results, counters = process_root(root_dir)
            all_results.extend(results)
            per_root_summary.append((root_dir, counters))

            print(
                f"[DONE] {root_dir} -> "
                f"OK={counters['OK']}, MISSING={counters['MISSING']}, INVALID={counters['INVALID']}, "
                f"DIFF={counters['DIFF']}, EXTRA={counters['EXTRA']}, ERROR={counters['ERROR']}"
            )
        except Exception as e:
            print(f"[ROOT ERROR] {root_dir} -> {e}")
            per_root_summary.append(
                (
                    root_dir,
                    {"OK": 0, "MISSING": 0, "INVALID": 0, "DIFF": 0, "EXTRA": 0, "ERROR": 1},
                )
            )
            all_results.append(
                CompareResult(
                    root_dir=root_dir,
                    rel_path=Path(),
                    lang="",
                    status="ERROR",
                    file_path=root_dir,
                    message=str(e),
                )
            )

    output_path = build_report(all_results, per_root_summary)

    print()
    print(f"Excel report created: {output_path}")


if __name__ == "__main__":
    main()